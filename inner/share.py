'''
Author: Mr.Car
Date: 2024-02-29 18:13:50
'''
import pandas as pd
import geopandas as gpd
import os
import yaml
import openpyxl
from openpyxl import load_workbook
from .load_config import csv_data, yaml_data, inner_xls_template_path, cfg_index
class XlsPosUtil:
    '''
    处理 excel 中的 position 字符串
    '''
    def __init__(self, position=""):
        if position:
            self.position = position
            self.start_row, self.start_col = int(position[1:]), ord(position[0].upper()) - 64 # B4 转为 列标数字
    def get_row_col(self, position):
        row, col = int(position[1:]), ord(position[0].upper()) - 64 # B4 转为 列标数字
        return row, col

    def get_excel_position(self, row, col):
        '''
        openpyxl 没有直接提供转换函数，但可以使用 get_column_letter 函数来获取列的字母表示，然后拼接行号;
        4,1 -> A4        
        '''
        # openpyxl 没有直接提供转换函数，但可以使用 get_column_letter 函数
        # 来获取列的字母表示，然后拼接行号。
        col_letter = openpyxl.utils.get_column_letter(col)
        return f"{col_letter}{row}"

    def position_add_row(self, position, num):
        row, col = self.get_row_col(position)
        row += num
        return self.get_excel_position(row, col)

    def position_add_col(self, position, num):
        row, col = self.get_row_col(position)
        col += num
        return self.get_excel_position(row, col)

    def segment(self, result_list, start_position, seg_length, position_interval, horizon=True):
        '''
        处理数据过长时候的分行显示问题
        [1,2,3,4] -> [1,2,3][4]
        '''
        start_row, start_col = self.get_row_col(start_position)
        position_list = []
        group_result_list = []
        if seg_length == 0:
            position_list.append(start_position)
            group_result_list.append(result_list)
        else:
            groups = len(result_list) // seg_length
            for i in range(groups):
                group_result_list.append(result_list[seg_length*i : seg_length*(i+1)])
                if horizon:
                    start_row += i * position_interval
                else:
                    start_col += i * position_interval
                position_list.append(self.get_excel_position(start_row, start_col))
        return position_list, group_result_list

def fill_template(sheet, start_position, value_list, horizon=True, interval=0):
    start_row, start_col = int(start_position[1:]), ord(start_position[0].upper()) - 64 # B4 转为 列标数字
    for each in value_list:
        sheet.cell(row=start_row, column=start_col, value=each)
        if horizon:
            start_col += 1 + interval
        else:
            start_row += 1 + interval
    return sheet

def fill_title(sheet, title, position="A1"):
    sheet[position] = title 
    return sheet

def fill_value(sheet, value, position):
    sheet[position] = value
    return sheet

def get_sheet(wb, template):
    sheet_name = next(x for x in wb.sheetnames if template in x)
    sheet = wb[sheet_name]
    return sheet


def _get_weight_result(df,calc_field, weight_field):
    denominator = df[calc_field].sum()
    if denominator != 0:
        divider =  (df[calc_field] * df[weight_field]).copy().sum()
        result = format( divider / denominator,'.5f' )
    else:
        result = '/'
    return result

def _get_sorted_list(df, parent_field, sort_field):
    result_list = []
    if sort_field:
        temp_df = df[[parent_field, sort_field]].copy()
        sorted_temp_df = temp_df.sort_values(by=sort_field, ascending=True)
        result_list = sorted_temp_df[parent_field].unique().tolist()
    else:
        result_list = df[parent_field].unique().tolist()
    return result_list

def _get_devided_result(divider, denominator):
    result = 0
    if denominator != 0:
        result = divider / denominator
    return result

# 原 report.py 中的准备文件：
def read_and_prepare_file(file_pth, file_type="shp"):
    if file_type == 'csv':
        df = pd.read_csv(file_path)
    else:
        df = gpd.read_file(file_pth)
    return df

def get_cfg_params(cfg_index, name, result_type="variable"):
    if result_type == "variable":
        var_table = csv_data[cfg_index.loc[name]["var"]]
    if result_type == "file_pth":
        var_table = os.path.join(folder_path, cfg_index.loc[name]["var"])
    rule_table = csv_data[cfg_index.loc[name]["rule"]]
    field = cfg_index.loc[name]["field"]
    target_field = cfg_index.loc[name]["target_field"]
    title = cfg_index.loc[name]["title"]
    nan_filler = cfg_index.loc[name]["nan_filler"]
    return var_table, rule_table, field, target_field, title, nan_filler

def get_wb(xls_file_path):
    wb = load_workbook(xls_file_path)    
    return wb

def fill_column_title(var_table, sheet):
    '''
    填充标题列
    '''
    for name, row in var_table.iterrows():
        locate_position = row["locate_position"]
        title = row["title"]
        if not pd.isna(locate_position) and not pd.isna(title):
            sheet = fill_value(sheet, title, locate_position)
    return sheet

def deal_none(df, field, rule_table, result_type="ss"):
    '''
    输入样点数据 输出df(除了空值之外的) 输出df(空值)
    '''
    ss = df[field]
    rule_str = rule_table.loc["none"]["value"]
    ss_none = ss[ss.apply(lambda x: eval(rule_str, {"x": x}))]
    df_none = df[ss.apply(lambda x: eval(rule_str, {"x": x}))]
    ss_not_none = ss[ss.apply(lambda x: not eval(rule_str, {"x": x}))]
    df_not_none = df[ss.apply(lambda x: not eval(rule_str, {"x": x}))]
    if result_type == "ss":
        return ss_none, ss_not_none
    elif result_type == "df":
        return df_none, df_not_none

def save_xls(wb, xls_file_path):
    wb.save(xls_file_path)
    return

def prepare_cfg(origin_file_pth, cfg_name, parent_field, sort_field=None):
    '''
    准备好文件
    '''
    with alive_bar(1, title="生成配置文件:") as bar:
        TRSX_112.prepare(origin_file_pth, folder_path, cfg_name, parent_field, sort_field)
        bar()
    return "Done!"